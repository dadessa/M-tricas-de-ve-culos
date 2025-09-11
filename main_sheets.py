import os
import sys
# DON'T CHANGE THIS !!!
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import json
import requests
import csv
from io import StringIO
from datetime import datetime
from openpyxl import load_workbook
from flask import Flask, send_from_directory, jsonify, request
from flask_cors import CORS

app = Flask(__name__, static_folder=os.path.join(os.path.dirname(__file__), 'static'))
app.config['SECRET_KEY'] = 'asdf#FGSgvasgf$5$WGT'
CORS(app)

# ID da planilha do Google Sheets
GOOGLE_SHEETS_ID = '17TnGB6NpsziDec4fPH-d0TCQwk2LN0BAv6yjmIpyZnI'
GOOGLE_SHEETS_URL = f'https://docs.google.com/spreadsheets/d/{GOOGLE_SHEETS_ID}/export?format=csv&gid=1225239898'

def load_google_sheets_data():
    """Carrega dados diretamente do Google Sheets"""
    try:
        print("Carregando dados do Google Sheets...")
        
        # Fazer requisição para obter dados CSV
        response = requests.get(GOOGLE_SHEETS_URL, timeout=30)
        response.raise_for_status()
        
        # Processar CSV
        csv_data = StringIO(response.text)
        reader = csv.DictReader(csv_data)
        
        data = []
        for row in reader:
            # Processar cada linha
            processed_row = process_sheets_row(row)
            if processed_row:
                data.append(processed_row)
        
        print(f"Dados carregados com sucesso: {len(data)} registros")
        return data
        
    except Exception as e:
        print(f"Erro ao carregar dados do Google Sheets: {e}")
        # Fallback para arquivo local se houver erro
        return load_excel_data()

def process_sheets_row(row):
    """Processa uma linha dos dados do Google Sheets"""
    try:
        # Mapear colunas para nomes padronizados
        nome = row.get('Nome do veículo.', '').strip()
        if not nome:
            return None
            
        # Extrair cidade
        cidade = extract_city_from_sheets(row)
        
        # Extrair categoria baseada nas visualizações
        categoria = extract_category_from_sheets(row)
        
        # Extrair status baseado nos campos Expediente e Cookies
        status = extract_status_from_sheets(row)
        
        # Extrair visualizações
        views_agosto = clean_number(row.get('Total de Visualizações Agosto', '0'))
        views_julho = clean_number(row.get('Total de Visualizações Julho', '0'))
        views_junho = clean_number(row.get('Total de visualizações Junho', '0'))
        
        processed_row = {
            'Nome do veículo.\n': nome,
            'Cidade': cidade,
            'Categoria': categoria,
            'Status': status,
            'Total de Vizualizações Agosto': views_agosto,
            'Total de Visualizações Julho': views_julho,
            'Total de vizualizações Junho': views_junho,
            'Expediente': row.get('Expediente', ''),
            'Cookies': row.get('Cookies', ''),
            'Endereço no site': row.get('Endereço no site', '')
        }
        
        return processed_row
        
    except Exception as e:
        print(f"Erro ao processar linha: {e}")
        return None

def clean_number(value):
    """Limpa e converte números"""
    try:
        if not value:
            return 0
        # Remove pontos e vírgulas, converte para int
        clean_value = str(value).replace('.', '').replace(',', '').replace(' ', '')
        return int(clean_value) if clean_value.isdigit() else 0
    except:
        return 0

def extract_city_from_sheets(row):
    """Extrai cidade dos dados do Google Sheets"""
    # Lista de possíveis colunas de cidade
    city_columns = ['Cidade', 'cidade', 'City', 'CIDADE']
    
    for col in city_columns:
        if col in row and row[col]:
            return row[col].strip()
    
    # Se não encontrar, usar lógica baseada no nome do veículo ou outros campos
    nome = row.get('Nome do veículo.', '').lower()
    
    # Mapear alguns nomes conhecidos para cidades
    city_mapping = {
        'maceio': 'Maceió',
        'arapiraca': 'Arapiraca',
        'penedo': 'Penedo',
        'delmiro': 'Delmiro Gouveia',
        'palmeira': 'Palmeira dos Índios',
        'brasilia': 'Brasília',
        'manaus': 'Manaus',
        'salvador': 'Salvador',
        'recife': 'Recife'
    }
    
    for key, city in city_mapping.items():
        if key in nome:
            return city
    
    # Default para Maceió se não conseguir determinar
    return 'Maceió'

def extract_category_from_sheets(row):
    """Extrai categoria baseada nas visualizações"""
    try:
        views_agosto = clean_number(row.get('Total de Visualizações Agosto', '0'))
        
        if views_agosto > 80000:
            return 'Mais de 80k'
        elif views_agosto > 50000:
            return '50k a 80k'
        elif views_agosto > 40000:
            return '40k a 50k'
        elif views_agosto > 30000:
            return '30k a 40k'
        elif views_agosto > 20000:
            return '20k a 30k'
        elif views_agosto > 10000:
            return '10k a 20k'
        else:
            return 'Menores 10k'
    except:
        return 'N/A'

def extract_status_from_sheets(row):
    """Extrai status baseado nos campos Expediente e Cookies"""
    expediente = row.get('Expediente', '').lower()
    cookies = row.get('Cookies', '').lower()
    endereco = row.get('Endereço no site', '').lower()
    
    # Lógica para determinar status
    has_expediente = 'possui' in expediente and 'não possui' not in expediente
    has_cookies = 'possui' in cookies and 'não possui' not in cookies
    has_endereco = 'possui' in endereco and 'não possui' not in endereco
    
    approved_count = sum([has_expediente, has_cookies, has_endereco])
    
    if approved_count >= 2:
        return 'APROVADO'
    elif approved_count == 1:
        return 'APROVADO PARCIAL'
    else:
        return 'REPROVADO'

# Carregar e processar dados do Excel (fallback)
def load_excel_data():
    try:
        excel_path = os.path.join(os.path.dirname(__file__), 'Recadastramento(respostas)(2).xlsx')
        wb = load_workbook(excel_path)
        ws = wb.active
        
        headers = []
        for cell in ws[1]:
            headers.append(cell.value if cell.value else '')
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = value
            data.append(row_data)
        
        return data
    except Exception as e:
        print(f"Erro ao carregar Excel: {e}")
        return []

# Carregar dados (prioriza Google Sheets)
try:
    data = load_google_sheets_data()
except:
    data = load_excel_data()

@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/dashboard/')
def dashboard():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/api/data')
def get_data():
    """Retorna todos os dados (sempre atualizado do Google Sheets)"""
    try:
        # Sempre buscar dados atualizados do Google Sheets
        fresh_data = load_google_sheets_data()
        return jsonify(fresh_data)
    except:
        return jsonify(data)

@app.route('/api/stats')
def get_stats():
    """Retorna estatísticas dos dados"""
    try:
        # Buscar dados atualizados
        fresh_data = load_google_sheets_data()
        
        total = len(fresh_data)
        aprovados = len([d for d in fresh_data if d.get('Status') == 'APROVADO'])
        reprovados = len([d for d in fresh_data if d.get('Status') == 'REPROVADO'])
        cidades = len(set([d.get('Cidade', 'N/A') for d in fresh_data if d.get('Cidade')]))
        
        return jsonify({
            'total': total,
            'aprovados': aprovados,
            'reprovados': reprovados,
            'cidades': cidades
        })
    except:
        total = len(data)
        aprovados = len([d for d in data if d.get('Status') == 'APROVADO'])
        reprovados = len([d for d in data if d.get('Status') == 'REPROVADO'])
        cidades = len(set([d.get('Cidade', 'N/A') for d in data if d.get('Cidade')]))
        
        return jsonify({
            'total': total,
            'aprovados': aprovados,
            'reprovados': reprovados,
            'cidades': cidades
        })

@app.route('/api/filter')
def filter_data():
    """Filtra dados baseado nos parâmetros"""
    try:
        # Buscar dados atualizados
        fresh_data = load_google_sheets_data()
    except:
        fresh_data = data
    
    cidade = request.args.get('cidade', '')
    status = request.args.get('status', '')
    categoria = request.args.get('categoria', '')
    
    filtered_data = fresh_data
    
    if cidade and cidade != 'Todas':
        filtered_data = [d for d in filtered_data if d.get('Cidade') == cidade]
    
    if status and status != 'Todos':
        filtered_data = [d for d in filtered_data if d.get('Status') == status]
    
    if categoria and categoria != 'Todas':
        filtered_data = [d for d in filtered_data if d.get('Categoria') == categoria]
    
    return jsonify(filtered_data)

@app.route('/api/search')
def search_data():
    """Busca dados por nome"""
    try:
        # Buscar dados atualizados
        fresh_data = load_google_sheets_data()
    except:
        fresh_data = data
    
    query = request.args.get('q', '').lower()
    
    if not query or len(query) < 2:
        return jsonify([])
    
    results = []
    for item in fresh_data:
        nome = item.get('Nome do veículo.\n', '').lower()
        if query in nome:
            results.append(item)
            if len(results) >= 20:  # Limitar resultados
                break
    
    return jsonify(results)

@app.route('/api/top-views-august')
def get_top_views_august():
    """Retorna top 10 visualizações de agosto"""
    try:
        # Buscar dados atualizados
        fresh_data = load_google_sheets_data()
    except:
        fresh_data = data
    
    # Filtrar e ordenar por visualizações de agosto
    valid_data = []
    for item in fresh_data:
        try:
            views = int(item.get('Total de Vizualizações Agosto', 0))
            if views > 0:
                valid_data.append({
                    'nome': item.get('Nome do veículo.\n', ''),
                    'views': views
                })
        except:
            continue
    
    # Ordenar por visualizações (decrescente) e pegar top 10
    top_10 = sorted(valid_data, key=lambda x: x['views'], reverse=True)[:10]
    
    return jsonify(top_10)

@app.route('/api/refresh')
def refresh_data():
    """Força atualização dos dados do Google Sheets"""
    try:
        global data
        data = load_google_sheets_data()
        return jsonify({'status': 'success', 'message': 'Dados atualizados com sucesso', 'count': len(data)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)

